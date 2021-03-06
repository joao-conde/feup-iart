from macros import *
from random import *
import openpyxl as sheets


"""
    Export the schedule to excel spreadsheet.
"""
def export_to_spreadsheet(wb_path, individual):
    wb = sheets.load_workbook('../files/template.xlsx')
    ws = wb.active

    center = sheets.styles.Alignment(horizontal='center', vertical='center')

    # Place room headers.
    for room_i in range(NUMBER_OF_ROOMS):
        for day_i in range(3):
            c = SHEET_COL_START + day_i * (NUMBER_OF_ROOMS + 1) + room_i
            ws.cell(row=3, column=c, value=f'Room #{room_i + 1}').alignment = center

    # Place day headers.
    for day_i in range(3):
        c = SHEET_COL_START + day_i * (NUMBER_OF_ROOMS + 1)
        ws.merge_cells(start_row=2, start_column=c, end_row=2, end_column=c+NUMBER_OF_ROOMS - 1)
        ws.cell(row=2, column=c, value=f'Day #{day_i + 1}').alignment = center
        
    # Write to spreadsheet the papers.
    for talk in individual:
        c = SHEET_COL_START + (talk['day'] - 1) * (NUMBER_OF_ROOMS + 1) + (talk['room'] - 1)
        r = SHEET_ROW_START + talk['time']

        for dur in range(talk['paper'].duration // 10):
            cell_value = str(talk['paper'].id) + ' - ' + talk['paper'].title
            ws.cell(row=r + dur, column=c, value=cell_value)

    wb.save(wb_path + '.xlsx')


"""
    Generates an integer in range [lb,up] not in the invalid set.
"""
def generate_except(lb, up, invalid):
    gen = randint(lb,up)
    
    while gen in invalid:
        gen = randint(lb, up)

    return gen

    