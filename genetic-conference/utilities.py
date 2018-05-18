from macros import *
import openpyxl as sheets


"""
"""
def export_to_spreadsheet(wb_path, individual):
    wb = sheets.load_workbook('template.xlsx')
    ws = wb.active

    center = sheets.styles.Alignment(horizontal='center', vertical='center')

    # Place room headers.
    for room_i in range(NUMBER_OF_ROOMS):
        for day_i in range(3):
            c = SHEET_COL_START + day_i * (NUMBER_OF_ROOMS + 1) + room_i
            ws.cell(row=3, column=c, value=f'Room #{room_i + 1}').alignment = center

    # Place day headers.
    for day_i in range(3):
        c = SHEET_COL_START + day_i * (NUMBER_OF_ROOMS + 1)
        ws.merge_cells(start_row=2, start_column=c, end_row=2, end_column=c+NUMBER_OF_ROOMS - 1)
        ws.cell(row=2, column=c, value=f'Day #{day_i + 1}').alignment = center
        
    # Write to spreadsheet the papers.
    for talk in individual:
        c = SHEET_COL_START + (talk['day'] - 1) * (NUMBER_OF_ROOMS + 1) + (talk['room'] - 1)
        r = SHEET_ROW_START + talk['time']

        ws.cell(row=r , column=c , value=talk['paper'].id)

    wb.save('results.xlsx')



def print_conference(conference, fitness = -1):
    
    if fitness != -1: print("FITNESS", fitness, '\n')
    
    for presentation in conference:
        print(presentation['paper'], " --> room:", presentation['room'], "; day:", presentation['day'], "; time:", presentation['time'])
